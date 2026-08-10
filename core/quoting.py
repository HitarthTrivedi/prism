"""
Prism — working out the price
─────────────────────────────
Two ways a quotation gets its numbers, because small manufacturers work both
ways and a tool that only understood one of them would not fit anybody:

  **Catalogue.**  The item is on their rate list. Look up the row, apply the
  quantity slab, add tax. Traders, dealers and stockists live here.

  **Cost sheet.**  The item is made to the customer's drawing, so no rate
  exists — nobody has made this exact part before. What the owner does in his
  head is: material weight × rate per kg, plus the process charges, plus
  margin. That is a cost sheet, and it is how every job shop prices.

────────────────────────────────────────────────────────────────────────────
No AI touches a number in this file
────────────────────────────────────────────────────────────────────────────
Every figure here is computed in Python, in Decimal, and a language model is
only ever asked to write the sentences *around* the numbers it is handed. A
model doing arithmetic is approximately right, and approximately right on a
rate that goes to a customer is a loss, an argument, or both.

Decimal rather than float for the same reason. 0.1 + 0.2 is not 0.3 in binary
floating point, GST at 18% of ₹1,42,500 has to come out the same as the
customer's own calculator, and "the software rounds differently from Tally" is
not a conversation worth having.
"""
from __future__ import annotations

import csv
import math
import os
import re
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP

PAISE = Decimal("0.01")


def rupees(value) -> Decimal:
    """Round to paise the way an invoice does — half up, never banker's.

    Python's default rounds .5 to the nearest even number, so ₹0.125 becomes
    ₹0.12. Every accounting system in India rounds it to ₹0.13, and a
    one-paisa disagreement across a hundred lines is exactly the kind of thing
    that makes somebody stop trusting the software.
    """
    if not isinstance(value, Decimal):
        value = Decimal(str(value or 0))
    return value.quantize(PAISE, rounding=ROUND_HALF_UP)


# A money cell as an Indian office actually writes it. The whole string has to
# match: currency in front, the number, then at most one piece of decoration
# after it. Being strict is the point — a cell holding two numbers, or prose,
# is not a price, and picking the first digits out of it would be a guess
# wearing the clothes of an answer.
#
#   Rs.1,000/-   ₹ 1,42,500.00   28.50/nos   1,000 nos   -500   142500
#
# The trailing "/-" is why this is a regex rather than a character strip. The
# first version deleted everything that was not a digit, a dot or a minus,
# which turned "Rs.1,000/-" into ".1000-" — unparseable, and so silently zero.
# Every order value typed the normal way summarised as ₹0, and a rate written
# "28.50/-" quoted at nothing.
_MONEY_CELL = re.compile(r"""
    ^\s*
    (?:rs\.?|inr|₹|₹)?\s*          # optional currency in front
    (?P<number>-?\d[\d,]*(?:\.\d+)?)    # the number itself
    \s*(?:/-|/=)?                       # the Indian "and no paise" ending
    \s*(?:/\s*)?[a-z%.]{0,12}           # an optional trailing unit: /nos, kg
    \s*$
""", re.IGNORECASE | re.VERBOSE)


def to_decimal(value, default: str = "0") -> Decimal:
    """Whatever was in the cell, as a Decimal. Unreadable cells give `default`.

    Deliberately returns the default rather than raising: one mistyped cell in
    a five-thousand-row rate list must not stop the whole file loading. The
    caller sees a zero, which is visible; an exception would lose the other
    4,999 rows.
    """
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    match = _MONEY_CELL.match(str(value or ""))
    if not match:
        return Decimal(default)
    try:
        return Decimal(match.group("number").replace(",", ""))
    except Exception:
        return Decimal(default)


# ── the rate list ─────────────────────────────────────────────────────────────

# Every small business names these columns differently and none of them will
# rename theirs to suit us. Recognising the obvious spellings is the difference
# between "upload your price list" and "reformat your price list first".
_COLUMN_ALIASES = {
    "code": ("code", "item code", "itemcode", "sku", "part no", "part number",
             "product code", "cat no", "catalogue no", "material code"),
    "description": ("description", "item", "item name", "product", "particulars",
                    "product name", "material", "details", "specification"),
    "unit": ("unit", "uom", "u.o.m", "per", "units"),
    "rate": ("rate", "price", "unit rate", "unit price", "rate per unit",
             "amount", "mrp", "list price", "basic rate", "rate/unit"),
    "hsn": ("hsn", "hsn code", "hsn/sac", "sac"),
    "moq": ("moq", "min qty", "minimum qty", "minimum order", "min order qty"),
    "group": ("group", "category", "type", "series", "family"),
}


@dataclass
class RateItem:
    code: str = ""
    description: str = ""
    unit: str = "nos"
    rate: Decimal = Decimal(0)
    hsn: str = ""
    moq: Decimal = Decimal(0)
    group: str = ""
    # (minimum quantity, rate at that quantity), biggest break last. Read from
    # extra "rate @ 100" style columns if the file has them.
    slabs: list[tuple[Decimal, Decimal]] = field(default_factory=list)

    def rate_for(self, quantity) -> Decimal:
        """The rate that applies at this quantity — the deepest slab reached."""
        quantity = to_decimal(quantity)
        rate = self.rate
        for minimum, slab_rate in sorted(self.slabs, key=lambda s: s[0]):
            if quantity >= minimum:
                rate = slab_rate
        return rate

    @property
    def label(self) -> str:
        return f"{self.code} — {self.description}".strip(" —") or self.description


_SLAB_HEADER = re.compile(r"(?:rate|price)\s*(?:@|at|above|for|>=?)\s*([\d,]+)", re.I)


def _canonical(header: str) -> str | None:
    h = (header or "").strip().lower().replace("_", " ")
    h = re.sub(r"\s+", " ", h).strip(" .:")
    for key, aliases in _COLUMN_ALIASES.items():
        if h in aliases:
            return key
    return None


def _rows_from_csv(path: str) -> list[list[str]]:
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [[(c or "").strip() for c in row] for row in csv.reader(f, dialect)]


def _rows_from_xlsx(path: str) -> list[list[str]]:
    """Read the first sheet of an .xlsx, if openpyxl is installed.

    Optional on purpose. Most rate lists are Excel files, so supporting them
    removes a real piece of friction — but a build without openpyxl must still
    work rather than refuse, so the caller is told to save as CSV instead.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RateFileError(
            "Prism can't open Excel files in this build. In Excel choose "
            "File → Save As → CSV, and upload that instead — the prices are "
            "identical.") from e
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        rows = []
        for raw in sheet.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in raw])
        return rows
    finally:
        book.close()


class RateFileError(Exception):
    """The rate list could not be understood — always with a fix in the text."""


def load_rates(path: str) -> list[RateItem]:
    """Read a price list into RateItems.

    Finds the header row rather than assuming row 1, because real price lists
    open with a company name, an address and a blank line before the table
    starts. Anything below the header that has a description and a number is
    an item; anything else is a section heading or a blank and is skipped.
    """
    if not os.path.exists(path):
        raise RateFileError(f"No rate list at {path}.")
    extension = os.path.splitext(path)[1].lower()
    if extension in (".xlsx", ".xlsm"):
        rows = _rows_from_xlsx(path)
    elif extension == ".xls":
        raise RateFileError(
            "That is the old .xls format, which Prism can't read. Open it in "
            "Excel and use File → Save As → CSV.")
    else:
        rows = _rows_from_csv(path)

    header_index, mapping = _find_header(rows)
    if header_index is None:
        raise RateFileError(
            "Prism couldn't find the price columns in that file. It needs a "
            "row of headings with at least a description column and a rate "
            "column — for example: Code, Description, Unit, Rate.")

    slabs = _slab_columns(rows[header_index])
    items = []
    for raw in rows[header_index + 1:]:
        item = _item_from(raw, mapping, slabs)
        if item is not None:
            items.append(item)
    if not items:
        raise RateFileError(
            "That file has the right headings but no priced rows underneath "
            "them. Check that the rate column has numbers in it.")
    return items


def _find_header(rows: list[list[str]]) -> tuple[int | None, dict]:
    """The first row that names a description and a rate. Scans a reasonable
    way down, because a price list often has a letterhead above the table."""
    for index, row in enumerate(rows[:40]):
        mapping = {}
        for position, cell in enumerate(row):
            key = _canonical(cell)
            if key and key not in mapping:
                mapping[key] = position
        if "rate" in mapping and ("description" in mapping or "code" in mapping):
            return index, mapping
    return None, {}


def _slab_columns(header: list[str]) -> list[tuple[int, Decimal]]:
    """Extra columns like "Rate @ 100" become quantity breaks."""
    out = []
    for position, cell in enumerate(header):
        match = _SLAB_HEADER.search(cell or "")
        if match:
            out.append((position, to_decimal(match.group(1))))
    return out


def _item_from(row: list[str], mapping: dict,
               slabs: list[tuple[int, Decimal]]) -> RateItem | None:
    def cell(key: str) -> str:
        position = mapping.get(key)
        return (row[position].strip() if position is not None and position < len(row)
                else "")

    description = cell("description")
    code = cell("code")
    if not description and not code:
        return None
    rate_text = cell("rate")
    # A row with no number in the rate column is a section heading
    # ("COMPRESSION SPRINGS"), not a product.
    if not re.search(r"\d", rate_text):
        return None

    item = RateItem(
        code=code,
        description=description or code,
        unit=cell("unit") or "nos",
        rate=to_decimal(rate_text),
        hsn=cell("hsn"),
        moq=to_decimal(cell("moq")),
        group=cell("group"),
    )
    for position, minimum in slabs:
        if position < len(row) and re.search(r"\d", row[position] or ""):
            item.slabs.append((minimum, to_decimal(row[position])))
    return item


# ── matching what the customer asked for to a row on the list ────────────────

_STOPWORDS = {"the", "a", "an", "of", "for", "and", "or", "with", "in", "on",
              "to", "please", "kindly", "need", "require", "required", "want",
              "send", "quote", "quotation", "rate", "rates", "price", "prices",
              "pcs", "nos", "no", "qty", "quantity", "dear", "sir", "madam",
              "regards", "thanks", "thank", "you", "we", "our", "us", "is",
              "are", "be", "as", "per", "your"}

_TOKEN = re.compile(r"[a-z]+|\d+(?:\.\d+)?")


def tokens(text: str) -> list[str]:
    """Words and numbers, lowercased, with the polite noise removed.

    Numbers are kept as tokens because in this trade they carry the meaning:
    "2 mm wire 25 od" is a specification, and dropping the digits would leave
    "mm wire od", which matches every spring ever made.
    """
    return [t for t in _TOKEN.findall((text or "").lower())
            if t not in _STOPWORDS and len(t) > 1]


@dataclass
class Match:
    item: RateItem
    score: float
    matched: list[str] = field(default_factory=list)

    @property
    def reason(self) -> str:
        """Why this row was picked, in the owner's words. Shown next to every
        suggestion — a match nobody can check is a match nobody should trust."""
        return ("matched on " + ", ".join(self.matched)) if self.matched else "closest wording"


def match_item(query: str, items: list[RateItem], limit: int = 5) -> list[Match]:
    """Best candidate rows for a free-text request, best first.

    Scored by how many of the request's words appear in the row, weighted so
    that a rare word counts for more than a common one — "cadmium" identifies a
    product, "spring" does not, on a spring maker's price list. Digits are
    weighted up again on top of that, because they are the specification.

    This returns candidates, never a decision. The caller shows them and the
    owner picks, which is the only honest thing to do with a fuzzy match on
    something that becomes a price.
    """
    wanted = tokens(query)
    if not wanted or not items:
        return []

    # How many rows each word appears in — a word in every row tells us nothing.
    frequency: dict[str, int] = {}
    row_tokens = []
    for item in items:
        item_tokens = set(tokens(f"{item.code} {item.description} {item.group}"))
        row_tokens.append(item_tokens)
        for token in item_tokens:
            frequency[token] = frequency.get(token, 0) + 1

    total = len(items)
    results = []
    for item, item_tokens in zip(items, row_tokens):
        score, matched = 0.0, []
        for token in set(wanted):
            if token not in item_tokens:
                continue
            weight = math.log(1 + total / (1 + frequency.get(token, 0)))
            if token.replace(".", "").isdigit():
                weight *= 2.0
            score += weight
            matched.append(token)
        if score > 0:
            # Long descriptions would otherwise win just by containing more
            # words; divide by their length so a precise short row can win.
            score /= math.sqrt(len(item_tokens) or 1)
            results.append(Match(item, round(score, 4), sorted(matched)))

    results.sort(key=lambda m: -m.score)
    return results[:limit]


def is_confident(matches: list[Match], margin: float = 1.6) -> bool:
    """True when the best row is clearly ahead of the runner-up.

    The gate for sending a quotation without a person looking. One good match
    is a match; two similar ones mean the request was ambiguous, and guessing
    between them is precisely the mistake that costs money.
    """
    if not matches:
        return False
    if len(matches) == 1:
        return matches[0].score > 0
    return matches[0].score >= matches[1].score * margin


# ── the cost sheet, for made-to-drawing work ─────────────────────────────────

# How each line of a cost sheet is charged. These four cover every job-shop
# cost sheet we have seen: material by weight, processes per piece, setup once
# per order, and overheads or margin as a percentage.
PER_KG = "per_kg"
PER_PIECE = "per_piece"
PER_LOT = "per_lot"
PERCENT = "percent"
BASES = (PER_KG, PER_PIECE, PER_LOT, PERCENT)


@dataclass
class CostLine:
    name: str
    basis: str = PER_PIECE
    rate: Decimal = Decimal(0)

    def amount(self, *, weight_kg: Decimal, quantity: Decimal,
               running: Decimal) -> Decimal:
        if self.basis == PER_KG:
            return self.rate * weight_kg * quantity
        if self.basis == PER_PIECE:
            return self.rate * quantity
        if self.basis == PER_LOT:
            return self.rate
        if self.basis == PERCENT:
            return running * self.rate / Decimal(100)
        return Decimal(0)


def load_cost_lines(path: str) -> list[CostLine]:
    """Read the owner's own cost sheet: name, basis, rate.

    Their sheet, their line names, their rates. Prism supplies the arithmetic
    and nothing else — the moment we ship a built-in idea of what a spring
    costs, we are wrong for every shop but one.
    """
    if not os.path.exists(path):
        raise RateFileError(f"No cost sheet at {path}.")
    rows = (_rows_from_xlsx(path) if path.lower().endswith((".xlsx", ".xlsm"))
            else _rows_from_csv(path))
    lines = []
    for row in rows:
        if len(row) < 3:
            continue
        name, basis, rate = row[0].strip(), row[1].strip().lower(), row[2]
        basis = basis.replace(" ", "_").replace("/", "_")
        if basis in ("per_kg", "kg", "by_weight"):
            basis = PER_KG
        elif basis in ("per_piece", "piece", "pc", "per_pc", "each", "nos"):
            basis = PER_PIECE
        elif basis in ("per_lot", "lot", "setup", "fixed", "once", "per_order"):
            basis = PER_LOT
        elif basis in ("percent", "percentage", "%", "pct"):
            basis = PERCENT
        else:
            continue           # a heading row, or something we don't understand
        if not name or not re.search(r"\d", str(rate)):
            continue
        lines.append(CostLine(name, basis, to_decimal(rate)))
    if not lines:
        raise RateFileError(
            "Prism couldn't read any cost lines from that file. Each row needs "
            "three things: the name of the charge, how it is charged "
            "(per_kg, per_piece, per_lot or percent), and the rate. "
            "For example:  Wire, per_kg, 95")
    return lines


@dataclass
class CostBreakdown:
    lines: list[tuple[str, Decimal]] = field(default_factory=list)
    total: Decimal = Decimal(0)
    quantity: Decimal = Decimal(1)

    @property
    def per_piece(self) -> Decimal:
        return rupees(self.total / self.quantity) if self.quantity else Decimal(0)


def cost_sheet(lines: list[CostLine], *, weight_kg, quantity) -> CostBreakdown:
    """Run the owner's cost sheet for one job. Shown, never hidden.

    Percentage lines apply to the running total above them, so the order of
    rows in their file is the order of their own working — which means the
    printout matches the way they already do it on paper.
    """
    weight_kg = to_decimal(weight_kg)
    quantity = to_decimal(quantity) or Decimal(1)
    out = CostBreakdown(quantity=quantity)
    running = Decimal(0)
    for line in lines:
        amount = line.amount(weight_kg=weight_kg, quantity=quantity, running=running)
        running += amount
        out.lines.append((line.name, rupees(amount)))
    out.total = rupees(running)
    return out


# ── weight helpers, for cost sheets charged by the kilogram ──────────────────
# Geometry, not opinion. Useful to anyone pricing by material weight — wire,
# bar, sheet or tube — and the numbers can be checked against a scale.

DENSITY_KG_PER_M3 = {
    "steel": Decimal("7850"), "spring steel": Decimal("7850"),
    "stainless": Decimal("7930"), "ss": Decimal("7930"),
    "brass": Decimal("8500"), "copper": Decimal("8960"),
    "aluminium": Decimal("2700"), "aluminum": Decimal("2700"),
}


def density_for(material: str) -> Decimal:
    key = (material or "").strip().lower()
    for name, value in DENSITY_KG_PER_M3.items():
        if name in key:
            return value
    return DENSITY_KG_PER_M3["steel"]


def wire_weight_kg(diameter_mm, length_mm, material: str = "steel") -> Decimal:
    """Weight of a length of round wire or bar."""
    diameter = to_decimal(diameter_mm)
    length = to_decimal(length_mm)
    if diameter <= 0 or length <= 0:
        return Decimal(0)
    radius_m = diameter / Decimal(2) / Decimal(1000)
    length_m = length / Decimal(1000)
    volume = Decimal(str(math.pi)) * radius_m * radius_m * length_m
    return volume * density_for(material)


def coil_length_mm(mean_diameter_mm, total_coils) -> Decimal:
    """Wire needed for a helical coil — π × mean diameter × number of coils.

    Mean diameter is outside diameter minus one wire diameter, which is the
    part people get wrong when they estimate by hand.
    """
    return Decimal(str(math.pi)) * to_decimal(mean_diameter_mm) * to_decimal(total_coils)


def spring_wire_weight_kg(*, wire_dia_mm, outer_dia_mm, total_coils,
                          material: str = "spring steel") -> Decimal:
    """Weight of one helical spring, from the three numbers on the drawing."""
    wire = to_decimal(wire_dia_mm)
    mean = to_decimal(outer_dia_mm) - wire
    if mean <= 0:
        return Decimal(0)
    return wire_weight_kg(wire, coil_length_mm(mean, total_coils), material)


# ── terms, lines and the quotation itself ────────────────────────────────────

@dataclass
class Terms:
    gst_percent: Decimal = Decimal(18)
    freight: Decimal = Decimal(0)
    discount_percent: Decimal = Decimal(0)
    validity_days: int = 15
    payment: str = "100% against proforma invoice"
    delivery: str = "2–3 weeks from receipt of confirmed order"
    currency: str = "INR"
    notes: str = ""


@dataclass
class QuoteLine:
    description: str
    quantity: Decimal
    unit: str = "nos"
    rate: Decimal = Decimal(0)
    hsn: str = ""
    # Where the rate came from — "rate list", "cost sheet", "entered by hand".
    # Printed on the internal copy so any figure can be traced later.
    basis: str = ""

    @property
    def amount(self) -> Decimal:
        return rupees(self.rate * self.quantity)


@dataclass
class Quotation:
    number: str = ""
    date: date | None = None
    customer: str = ""
    contact: str = ""
    email: str = ""
    inquiry_no: str = ""
    lines: list[QuoteLine] = field(default_factory=list)
    terms: Terms = field(default_factory=Terms)
    breakdown: CostBreakdown | None = None

    # ── the arithmetic, in one place so it cannot disagree with itself ──
    @property
    def subtotal(self) -> Decimal:
        return rupees(sum((l.amount for l in self.lines), Decimal(0)))

    @property
    def discount(self) -> Decimal:
        return rupees(self.subtotal * self.terms.discount_percent / Decimal(100))

    @property
    def taxable(self) -> Decimal:
        return rupees(self.subtotal - self.discount + rupees(self.terms.freight))

    @property
    def gst(self) -> Decimal:
        return rupees(self.taxable * self.terms.gst_percent / Decimal(100))

    @property
    def total(self) -> Decimal:
        return rupees(self.taxable + self.gst)

    @property
    def valid_until(self) -> date:
        return (self.date or date.today()) + timedelta(days=self.terms.validity_days)


def next_quote_number(rows: list[dict], prefix: str = "QTN",
                      when: date | None = None) -> str:
    """Next quotation number, counted off the register's Quotation no column.

    Shares register.next_number's rule so both books run on the same financial
    year and neither can drift from the other.
    """
    from . import register
    fake = [{"Inquiry no": r.get("Quotation no", "")} for r in rows]
    return register.next_number(fake, prefix, when)


def indian_currency(value: Decimal) -> str:
    """1,42,500.00 — lakh grouping, which is what the reader expects to see."""
    value = rupees(value)
    negative = value < 0
    whole, _, paise = f"{abs(value):.2f}".partition(".")
    if len(whole) > 3:
        head, tail = whole[:-3], whole[-3:]
        head = re.sub(r"(\d)(?=(\d\d)+$)", r"\1,", head)
        whole = f"{head},{tail}"
    return f"{'-' if negative else ''}{whole}.{paise}"


def render_text(quote: Quotation, company: str = "") -> str:
    """The quotation as plain text — the body of the mail, and the fallback
    when no template has been set up yet."""
    when = quote.date or date.today()
    out = []
    if company:
        out += [company, ""]
    out += [f"QUOTATION  {quote.number}",
            f"Date: {when.strftime('%d-%m-%Y')}"]
    if quote.inquiry_no:
        out.append(f"Against your inquiry: {quote.inquiry_no}")
    out += ["", f"To: {quote.customer}"]
    if quote.contact:
        out.append(f"Kind attention: {quote.contact}")
    out += ["", f"{'Sr':<3} {'Description':<40} {'Qty':>8} {'Unit':<6} "
                f"{'Rate':>12} {'Amount':>14}",
            "-" * 88]
    for index, line in enumerate(quote.lines, 1):
        out.append(f"{index:<3} {line.description[:40]:<40} "
                   f"{line.quantity:>8} {line.unit:<6} "
                   f"{indian_currency(line.rate):>12} "
                   f"{indian_currency(line.amount):>14}")
    out += ["-" * 88,
            f"{'Subtotal':>62} {indian_currency(quote.subtotal):>25}"]
    if quote.terms.discount_percent:
        out.append(f"{'Discount ' + str(quote.terms.discount_percent) + '%':>62} "
                   f"{'-' + indian_currency(quote.discount):>25}")
    if quote.terms.freight:
        out.append(f"{'Freight':>62} {indian_currency(quote.terms.freight):>25}")
    out += [f"{'GST ' + str(quote.terms.gst_percent) + '%':>62} "
            f"{indian_currency(quote.gst):>25}",
            f"{'TOTAL':>62} {indian_currency(quote.total):>25}", ""]
    out += ["Terms",
            f"  Payment  : {quote.terms.payment}",
            f"  Delivery : {quote.terms.delivery}",
            f"  Validity : {quote.terms.validity_days} days "
            f"(until {quote.valid_until.strftime('%d-%m-%Y')})"]
    if quote.terms.notes:
        out.append(f"  Note     : {quote.terms.notes}")
    return "\n".join(out)


def write_csv(quote: Quotation, path: str) -> str:
    """The line items as a CSV — for their records, and for Tally later."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Quotation no", quote.number])
        writer.writerow(["Date", (quote.date or date.today()).strftime("%d-%m-%Y")])
        writer.writerow(["Customer", quote.customer])
        writer.writerow(["Inquiry no", quote.inquiry_no])
        writer.writerow([])
        writer.writerow(["Sr", "Description", "HSN", "Quantity", "Unit",
                         "Rate", "Amount", "Rate source"])
        for index, line in enumerate(quote.lines, 1):
            writer.writerow([index, line.description, line.hsn,
                             f"{line.quantity}", line.unit,
                             f"{rupees(line.rate)}", f"{line.amount}", line.basis])
        writer.writerow([])
        writer.writerow(["", "", "", "", "", "Subtotal", f"{quote.subtotal}"])
        if quote.terms.discount_percent:
            writer.writerow(["", "", "", "", "",
                             f"Discount {quote.terms.discount_percent}%",
                             f"-{quote.discount}"])
        if quote.terms.freight:
            writer.writerow(["", "", "", "", "", "Freight",
                             f"{rupees(quote.terms.freight)}"])
        writer.writerow(["", "", "", "", "", f"GST {quote.terms.gst_percent}%",
                         f"{quote.gst}"])
        writer.writerow(["", "", "", "", "", "Total", f"{quote.total}"])
    return path


# ── the covering letter ───────────────────────────────────────────────────────

def covering_letter_prompt(quote: Quotation, inquiry_text: str = "",
                           signature: str = "") -> str:
    """Ask an AI to write the mail that carries the quotation.

    The numbers are handed over already formatted and the model is told, in as
    many words, not to compute or restate them. Its job is the two paragraphs
    of courtesy around the attachment — which is the one part of this workflow
    where a language model is genuinely better than a template.
    """
    return (
        "Write a short covering email to send with a quotation. Reply with "
        "NOTHING except the email, in exactly this format:\n\n"
        "SUBJECT: <one subject line>\n"
        "BODY:\n"
        "<the email body>\n\n"
        "Rules:\n"
        "- Do NOT include the price table, the rates or the total. The "
        "quotation is attached separately and repeating figures risks them "
        "disagreeing.\n"
        "- Do NOT calculate anything or mention any number that is not given "
        "to you below.\n"
        "- Three short paragraphs at most: thank them for the inquiry, say "
        "the quotation is attached and note the validity, and offer to "
        "discuss.\n"
        "- Plain business English as used in Indian manufacturing. Warm but "
        "brief. No marketing language, no exclamation marks.\n"
        "- No placeholders in square brackets. Everything you need is here.\n\n"
        f"Customer: {quote.customer}\n"
        f"Contact person: {quote.contact or 'not known — do not invent one'}\n"
        f"Quotation number: {quote.number}\n"
        f"Validity: {quote.terms.validity_days} days\n"
        f"Delivery: {quote.terms.delivery}\n"
        f"Payment terms: {quote.terms.payment}\n"
        f"What they asked for: {(inquiry_text or '').strip()[:800] or 'not stated'}\n"
        f"Sign off as: {signature or 'the sales team'}\n"
    )
