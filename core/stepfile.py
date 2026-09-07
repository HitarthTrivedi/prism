"""
Prism — /step: measure a 3D CAD model the way /gerber measures a PCB job
─────────────────────────────────────────────────────────────────────────
A customer sends a .STEP file of the product — a sheet-metal enclosure, a
plastic part — and the fab's estimator opens it in CAD and reads the
numbers off by hand: overall size, each part's size, wall thickness, every
hole. This module does that reading offline, on this machine, with the
same rules as the Gerber add-on:

  · No AI ever sees the STEP file. The customer's design is their product;
    only the measured numbers leave this module.
  · Everything shown is measured from the real geometry (OpenCascade, via
    the cadquery package) — never guessed, never generated.
  · Two decimal places everywhere a figure is shown or written.

What comes out, per part and for the assembly:

  · the formed (as-modelled) size L x W x H — NOTE: for bent sheet metal
    this is the finished part, not the unfolded flat sheet; the report
    says so, because a fab drawing often quotes the flat size;
  · a wall/sheet thickness ESTIMATE (2V/A — exact for a plain sheet, a
    little under for a part full of holes; the report calls it an
    estimate);
  · every hole, grouped by diameter with a count of distinct positions
    (a slot's two rounded ends count as two positions);
  · volume, surface area, and weight at common material densities —
    CRC steel for metal moulding, ABS/PP/PC/Nylon for plastic.

Deliverables written beside the terminal output, every one of them named
after the customer's own file — for Assem1.STEP:

  · Assem1 - dimensions.xlsx      — one row per part in the drawing's own
                                    style ("101.00 x 93.00 x 71.00 mm — 1
                                    nos") plus a hole table, readable by
                                    the estimator's Excel;
  · Assem1 - drawing sheet.html   — a drawing sheet: an isometric view of
    Assem1 - drawing sheet.png      every part with its caption and holes,
                                    like the hand-made sheet this replaces;
  · Assem1 - view <part>.svg      — the individual views.

The name is the whole point of the prefix: an estimator keeps ten jobs'
sheets in one folder, and "dimensions.xlsx" ten times over is ten files
nobody can tell apart. See names() and output_dir() — where the files go
is the caller's choice (the GUI asks the person), the default is a folder
per model under ~/Desktop/Prism Step.
"""
from __future__ import annotations

import datetime as _dt
import os
import re
from urllib.parse import quote as _quote

try:
    import cadquery as _cq
    from OCP.BRepAdaptor import BRepAdaptor_Surface
    from OCP.IFSelect import IFSelect_RetDone
    from OCP.STEPCAFControl import STEPCAFControl_Reader
    from OCP.TCollection import (TCollection_AsciiString,
                                 TCollection_ExtendedString)
    from OCP.TDataStd import TDataStd_Name
    from OCP.TDF import TDF_Label, TDF_LabelSequence
    from OCP.TDocStd import TDocStd_Document
    from OCP.TopoDS import TopoDS
    from OCP.XCAFDoc import XCAFDoc_DocumentTool
    HAVE_CAD = True
except Exception:                                   # pragma: no cover
    HAVE_CAD = False

try:
    import openpyxl
    HAVE_XLSX = True
except Exception:                                   # pragma: no cover
    HAVE_XLSX = False


class StepError(Exception):
    """A sentence for the user, not a stack trace."""


MODES = ("metal", "plastic")

# g/cm3. The estimator multiplies volume by one of these anyway; doing it
# here saves the calculator without pretending to know the exact alloy.
_DENSITY = {
    "metal": [("CRC steel", 7.85), ("aluminium", 2.70), ("SS 304", 8.00)],
    "plastic": [("ABS", 1.04), ("PP", 0.91), ("PC", 1.20), ("Nylon 6", 1.14)],
}

_EXTS = (".step", ".stp")

# Where a model's files go when nobody has said otherwise. The GUI asks the
# person and passes their answer as `root` to output_dir(); the terminal
# reads cfg["step_out_dir"] and falls back to this.
DEFAULT_OUT_ROOT = os.path.join(os.path.expanduser("~/Desktop"), "Prism Step")


def available() -> tuple[bool, str]:
    if not HAVE_CAD:
        return False, ("The STEP add-on needs the cadquery package "
                       "(pip install cadquery).")
    return True, ""


# ── where the files go, and what they are called ────────────────────────────

def stem_of(path_or_name: str) -> str:
    """The customer's own file name without its extension, made safe to use
    as part of a file name: 'Assem1.STEP' -> 'Assem1', '~/x/housing v2.stp'
    -> 'housing v2'. Case and spaces are kept — this is the name the person
    knows the job by, and 'assem1' is not what they called it."""
    base = os.path.splitext(os.path.basename(path_or_name or ""))[0]
    illegal = '<>:"/\\|?*'          # reserved on Windows; harmless elsewhere
    clean = "".join(c for c in base.strip() if c not in illegal and c.isprintable())
    clean = " ".join(clean.split())[:60].strip(" .")
    return clean or "model"


def _stem(report: dict) -> str:
    return report.get("stem") or stem_of(report.get("file", ""))


def names(report_or_stem) -> dict[str, str]:
    """Every deliverable's file name, each carrying the model's own name.

    One place, so the terminal, the GUI and the review page all agree on
    what a file is called and none of them writes a bare "dimensions.xlsx"
    again. Readable words with ' - ' between them, the same shape as the
    Artifacts folder's names (core.config._artifact_stem)."""
    stem = (report_or_stem if isinstance(report_or_stem, str)
            else _stem(report_or_stem))
    return {
        "xlsx":       f"{stem} - dimensions.xlsx",
        "html":       f"{stem} - drawing sheet.html",
        "svg":        f"{stem} - drawing sheet.svg",
        "png":        f"{stem} - drawing sheet.png",
        "review":     f"{stem} - change review.html",
        "modified":   f"{stem} - modified.step",
        "xlsx_after": f"{stem} - dimensions after change.xlsx",
        "png_after":  f"{stem} - drawing sheet after change.png",
        "after_dir":  f"{stem} - after change",
    }


def view_name(stem: str, part_name: str, index: int) -> str:
    """The SVG for one part: 'Assem1 - view top.svg'."""
    safe = re.sub(r"[^\w.-]", "_", part_name or "") or f"part{index}"
    return f"{stem} - view {safe}.svg"


def ai_sheet_name(stem: str, n: int, ext: str = ".png") -> str:
    """What /step-auto's returned drawing is saved as."""
    return f"{stem} - AI drawing sheet {n}{ext or '.png'}"


def output_dir(target: str, root: str = "") -> str:
    """The folder one model's files go into: <root>/<stem>.

    `root` is where the person said their STEP work should live (the GUI
    asks; the terminal has /step-folder); empty means DEFAULT_OUT_ROOT. A
    folder that already holds files gets a numbered sibling — 'Assem1 (2)',
    'Assem1 (3)' — the way Finder and Explorer do it, so measuring the same
    model twice never silently overwrites the first sheet. Nothing is
    created here; the writers make the folder when they write."""
    root = os.path.abspath(os.path.expanduser(root or DEFAULT_OUT_ROOT))
    stem = stem_of(target)
    first = os.path.join(root, stem)
    if not os.path.isdir(first) or not os.listdir(first):
        return first
    n = 2
    while True:
        cand = os.path.join(root, f"{stem} ({n})")
        if not os.path.isdir(cand) or not os.listdir(cand):
            return cand
        n += 1


# ── reading the file, names kept ─────────────────────────────────────────────

def _name_of(label) -> str:
    n = TDataStd_Name()
    if label.FindAttribute(TDataStd_Name.GetID_s(), n):
        return TCollection_AsciiString(n.Get()).ToCString()
    return ""


def load_parts(path: str) -> list[tuple[str, object]]:
    """[(part name, cadquery Shape)] with assembly placement applied.

    Through the XCAF reader rather than the plain importer, because the
    plain importer throws the part names away — and "top: 101.00 x 93.00"
    is worth a great deal more than "part 2: 101.00 x 93.00".
    """
    doc = TDocStd_Document(TCollection_ExtendedString("prism-step"))
    reader = STEPCAFControl_Reader()
    reader.SetNameMode(True)
    if reader.ReadFile(path) != IFSelect_RetDone:
        raise StepError(f"Could not read {os.path.basename(path)} — is it a "
                        "STEP file?")
    reader.Transfer(doc)
    tool = XCAFDoc_DocumentTool.ShapeTool_s(doc.Main())
    free = TDF_LabelSequence()
    tool.GetFreeShapes(free)

    parts: list[tuple[str, object]] = []
    for i in range(1, free.Length() + 1):
        root = free.Value(i)
        comps = TDF_LabelSequence()
        if tool.GetComponents_s(root, comps) and comps.Length():
            for j in range(1, comps.Length() + 1):
                comp = comps.Value(j)
                name = _name_of(comp)
                ref = TDF_Label()
                if tool.GetReferredShape_s(comp, ref):
                    name = _name_of(ref) or name
                parts.append((name or f"part {len(parts) + 1}",
                              _cq.Shape.cast(tool.GetShape_s(comp))))
        else:
            parts.append((_name_of(root) or f"part {len(parts) + 1}",
                          _cq.Shape.cast(tool.GetShape_s(root))))
    if not parts:
        raise StepError("The file holds no solid parts.")
    return parts


# ── measuring ────────────────────────────────────────────────────────────────

def _holes_of(shape) -> list[dict]:
    """Cylindrical faces grouped by diameter: [{dia_mm, positions}].

    Positions are distinct cylinder axes, so a hole counted from both its
    halves stays one hole. A slot's two rounded ends are two positions —
    honest, and the report says a slot reads that way.
    """
    groups: dict[float, set] = {}
    for f in shape.Faces():
        if f.geomType() != "CYLINDER":
            continue
        ad = BRepAdaptor_Surface(TopoDS.Face_s(f.wrapped))
        cyl = ad.Cylinder()
        dia = round(2 * cyl.Radius(), 2)
        p = cyl.Axis().Location()
        key = (round(p.X(), 1), round(p.Y(), 1), round(p.Z(), 1))
        groups.setdefault(dia, set()).add(key)
    return [{"dia_mm": dia, "count": len(axes)}
            for dia, axes in sorted(groups.items())]


def _measure(name: str, shape) -> dict:
    bb = shape.BoundingBox()
    dims = sorted((bb.xlen, bb.ylen, bb.zlen), reverse=True)
    volume = shape.Volume()             # mm3
    area = shape.Area()                 # mm2
    return {
        "name": name,
        "size_mm": tuple(round(d, 2) for d in dims),
        # Exact for a plain sheet; slightly under for a part full of holes.
        "thickness_mm": round(2 * volume / area, 2) if area else 0.0,
        "volume_cm3": round(volume / 1000.0, 2),
        "area_cm2": round(area / 100.0, 2),
        "holes": _holes_of(shape),
    }


def analyse(path: str, mode: str = "metal") -> dict:
    """Measure every part of a STEP file. Offline; nothing leaves here."""
    ok, why = available()
    if not ok:
        raise StepError(why)
    if mode not in MODES:
        raise StepError(f"Mode must be one of {MODES}, not {mode!r}.")
    path = os.path.abspath(os.path.expanduser(path))
    if not os.path.exists(path):
        raise StepError(f"No such file: {path}")
    if not path.lower().endswith(_EXTS):
        raise StepError("That is not a .step/.stp file.")

    named = load_parts(path)
    parts = [_measure(name, shape) for name, shape in named]
    parts.sort(key=lambda p: -p["volume_cm3"])

    compound = _cq.Compound.makeCompound([s for _, s in named])
    bb = compound.BoundingBox()
    overall = tuple(round(d, 2)
                    for d in sorted((bb.xlen, bb.ylen, bb.zlen), reverse=True))

    return {
        "file": os.path.basename(path),
        "path": path,
        # The name every deliverable is prefixed with — see names().
        "stem": stem_of(path),
        "mode": mode,
        "parts": parts,
        "overall_mm": overall,
        "_shapes": named,       # for the renderer; stripped before saving
        "warnings": [
            "Sizes are the FORMED part as modelled — for bent sheet metal "
            "this is the finished part, not the unfolded flat sheet a "
            "cutting list quotes.",
            "Thickness is an estimate (2 x volume / surface); a part full "
            "of holes reads a little under its nominal sheet.",
            "A slot's two rounded ends count as two hole positions.",
        ],
    }


# ── words and numbers out ────────────────────────────────────────────────────

def _weights(volume_cm3: float, mode: str) -> str:
    return " · ".join(f"{name} {volume_cm3 * dens:.2f} g"
                      for name, dens in _DENSITY[mode])


def _caption(part: dict, mode: str) -> str:
    L, W, H = part["size_mm"]
    stock = (f"t≈{part['thickness_mm']:.2f} mm sheet" if mode == "metal"
             else f"wall≈{part['thickness_mm']:.2f} mm")
    return f"{L:.2f} x {W:.2f} x {H:.2f} mm · {stock} — 1 nos"


def report_text(report: dict) -> str:
    lines = [f"STEP MEASUREMENT — {report['file']}   "
             f"({report['mode']} moulding)",
             "",
             f"  Assembly overall    {report['overall_mm'][0]:.2f} x "
             f"{report['overall_mm'][1]:.2f} x {report['overall_mm'][2]:.2f} mm",
             f"  Parts               {len(report['parts'])}",
             ""]
    for i, part in enumerate(report["parts"], 1):
        lines.append(f"  {i}) {part['name']}  —  {_caption(part, report['mode'])}")
        lines.append(f"     volume {part['volume_cm3']:.2f} cm3 · "
                     f"weight {_weights(part['volume_cm3'], report['mode'])}")
        if part["holes"]:
            lines.append("     holes  " + " · ".join(
                f"Ø{h['dia_mm']:g} x {h['count']}" for h in part["holes"]))
        lines.append("")
    for w in report["warnings"]:
        lines.append(f"  ! {w}")
    lines.append("")
    lines.append("  Measured offline from the geometry itself — the STEP "
                 "file never leaves this machine and no AI ever sees it.")
    return "\n".join(lines)


def write_xlsx(report: dict, path: str) -> str:
    """The estimator's sheet: one row per part, then every hole."""
    if not HAVE_XLSX:
        raise StepError("Writing Excel needs the openpyxl package.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts"
    bold = openpyxl.styles.Font(bold=True)

    ws.append([f"{report['file']} — {report['mode']} moulding — "
               f"measured {_dt.date.today().strftime('%d-%m-%Y')}"])
    ws["A1"].font = bold
    ws.append([])
    header = ["Sr", "Part", "L (mm)", "W (mm)", "H (mm)",
              "Thickness est (mm)", "Volume (cm3)",
              "Weight (g)", "Holes", "Line for the drawing"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = bold
    first = _DENSITY[report["mode"]][0]
    for i, part in enumerate(report["parts"], 1):
        L, W, H = part["size_mm"]
        holes = " · ".join(f"Ø{h['dia_mm']:g} x {h['count']}"
                           for h in part["holes"]) or "—"
        ws.append([i, part["name"], L, W, H, part["thickness_mm"],
                   part["volume_cm3"],
                   round(part["volume_cm3"] * first[1], 2),
                   holes, f"{i}) {_caption(part, report['mode'])}"])
    ws.append([])
    ws.append(["Assembly overall",
               f"{report['overall_mm'][0]:.2f} x {report['overall_mm'][1]:.2f}"
               f" x {report['overall_mm'][2]:.2f} mm"])
    ws.append([f"Weight column uses {first[0]} at {first[1]} g/cm3."])
    for w in report["warnings"]:
        ws.append([f"! {w}"])
    for col, width in zip("ABCDEFGHIJ", (4, 14, 10, 10, 10, 16, 12, 12, 40, 44)):
        ws.column_dimensions[col].width = width

    holes_ws = wb.create_sheet("Holes")
    holes_ws.append(["Part", "Diameter (mm)", "Positions"])
    for cell in holes_ws[1]:
        cell.font = bold
    for part in report["parts"]:
        for h in part["holes"]:
            holes_ws.append([part["name"], h["dia_mm"], h["count"]])
    for col, width in zip("ABC", (14, 14, 10)):
        holes_ws.column_dimensions[col].width = width

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb.save(path)
    return path


# ── the drawing sheet ────────────────────────────────────────────────────────

def _svg_of(shape, path: str) -> bool:
    """One part as a hidden-line isometric SVG. A view that cannot be drawn
    must not sink the measurement — the sheet shows the caption alone."""
    try:
        from cadquery.occ_impl.exporters.svg import getSVG
        svg = getSVG(shape, opts={"width": 420, "height": 340,
                                  "marginLeft": 24, "marginTop": 24,
                                  "showAxes": False, "showHidden": True,
                                  "projectionDir": (1, -1.2, 1)})
        with open(path, "w", encoding="utf-8") as f:
            f.write(svg)
        return True
    except Exception:                               # noqa: BLE001
        return False


# ── the dimensioned drawing sheet — drawn by Prism, no AI ──────────────────
#
# What the fab's estimator actually hands round is a dimension sheet: three
# orthographic views of each part with the overall sizes on dimension lines,
# a hole table, notes and a title block. /step-auto used to ask an image
# model for that from the measured numbers, which took minutes and came back
# looking right while quietly mis-stating a figure. Every line below is
# projected from the real geometry (OpenCascade hidden-line removal), every
# number is the measured one to two decimals, and it takes a second.

# The three standard views. `n` is the direction the camera looks FROM (the
# convention cadquery's exporter uses), `x` the axis that reads left-to-right
# on paper — fixed explicitly so Z is always up in the front and side views
# and Y is up in the top view, instead of whatever OpenCascade picks.
_VIEWS = (
    ("FRONT VIEW", (0, -1, 0), (1, 0, 0), ("x", "z")),
    ("TOP VIEW",   (0, 0, 1),  (1, 0, 0), ("x", "y")),
    ("SIDE VIEW",  (1, 0, 0),  (0, 1, 0), ("y", "z")),
)


def _project(shape, n, x=None) -> dict | None:
    """One orthographic view of a shape: {"visible": [svg path d…],
    "hidden": [...], "bb": (xmin, xmax, ymin, ymax)} in model millimetres,
    hidden lines removed the way a drawing office does it. None when the
    projection fails — a view that cannot be drawn must not sink the sheet."""
    try:
        from cadquery.occ_impl.exporters.svg import getPaths
        from cadquery.occ_impl.shapes import TOLERANCE, Shape, Compound
        from OCP.BRepLib import BRepLib
        from OCP.gp import gp_Ax2, gp_Dir, gp_Pnt
        from OCP.HLRAlgo import HLRAlgo_Projector
        from OCP.HLRBRep import HLRBRep_Algo, HLRBRep_HLRToShape

        hlr = HLRBRep_Algo()
        hlr.Add(shape.wrapped)
        cs = (gp_Ax2(gp_Pnt(), gp_Dir(*n), gp_Dir(*x)) if x
              else gp_Ax2(gp_Pnt(), gp_Dir(*n)))
        hlr.Projector(HLRAlgo_Projector(cs))
        hlr.Update()
        hlr.Hide()
        hs = HLRBRep_HLRToShape(hlr)
        visible = [c for c in (hs.VCompound(), hs.Rg1LineVCompound(),
                               hs.OutLineVCompound()) if not c.IsNull()]
        hidden = [c for c in (hs.HCompound(), hs.OutLineHCompound())
                  if not c.IsNull()]
        for el in visible + hidden:
            BRepLib.BuildCurves3d_s(el, TOLERANCE)
        visible = [Shape(c) for c in visible]
        hidden = [Shape(c) for c in hidden]
        hidden_paths, visible_paths = getPaths(visible, hidden)
        if not visible_paths and not hidden_paths:
            return None
        bb = Compound.makeCompound(hidden + visible).BoundingBox()
        return {"visible": visible_paths, "hidden": hidden_paths,
                "bb": (bb.xmin, bb.xmax, bb.ymin, bb.ymax)}
    except Exception:                               # noqa: BLE001
        return None


def _extents(shape) -> dict:
    bb = shape.BoundingBox()
    return {"x": bb.xlen, "y": bb.ylen, "z": bb.zlen}


class _Sheet:
    """An SVG page being drawn, in pixels. Small helpers so the layout code
    reads as a drawing, not as string concatenation."""

    W = 1240                    # A4 portrait at ~150 dpi; the height grows
    M = 44                      # page margin
    FONT = "-apple-system,'Segoe UI',Helvetica,Arial,sans-serif"

    def __init__(self):
        self.items: list[str] = []
        self.y = self.M         # the next free row

    def text(self, x, y, s, size=13, anchor="start", weight="normal",
             fill="#111", rotate=None):
        import html as _html
        tr = f' transform="rotate({rotate} {x} {y})"' if rotate else ""
        self.items.append(
            f'<text x="{x:.1f}" y="{y:.1f}" font-size="{size}" '
            f'text-anchor="{anchor}" font-weight="{weight}" fill="{fill}"'
            f'{tr}>{_html.escape(str(s))}</text>')

    def line(self, x1, y1, x2, y2, w=1, color="#111", dash=""):
        d = f' stroke-dasharray="{dash}"' if dash else ""
        self.items.append(
            f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" '
            f'stroke="{color}" stroke-width="{w}"{d}/>')

    def rect(self, x, y, w, h, stroke="#111", fill="none", sw=1):
        self.items.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{h:.1f}" '
            f'stroke="{stroke}" fill="{fill}" stroke-width="{sw}"/>')

    def arrow(self, x, y, dx, dy):
        """A filled arrowhead at (x, y) pointing along (dx, dy)."""
        import math
        L, Wd = 9.0, 3.2
        ang = math.atan2(dy, dx)
        bx, by = x - L * math.cos(ang), y - L * math.sin(ang)
        px, py = -math.sin(ang) * Wd, math.cos(ang) * Wd
        self.items.append(
            f'<polygon points="{x:.1f},{y:.1f} {bx + px:.1f},{by + py:.1f} '
            f'{bx - px:.1f},{by - py:.1f}" fill="#111"/>')

    def view(self, proj: dict, x0, y0, scale, w_px, h_px):
        """Draw a projection into the box at (x0, y0) of w_px x h_px, model
        units scaled by `scale`, centred. Returns the drawn extents
        (left, top, right, bottom) in px for the dimension lines."""
        xmin, xmax, ymin, ymax = proj["bb"]
        dw, dh = (xmax - xmin) * scale, (ymax - ymin) * scale
        left = x0 + (w_px - dw) / 2
        top = y0 + (h_px - dh) / 2
        # SVG y runs down the page; model y runs up. Flip about the box.
        tx = left - xmin * scale
        ty = top + ymax * scale
        g = (f'<g transform="translate({tx:.2f},{ty:.2f}) '
             f'scale({scale:.4f},{-scale:.4f})" fill="none" '
             f'vector-effect="non-scaling-stroke">')
        sw = 1.1 / scale
        g += (f'<g stroke="#8a8f94" stroke-width="{sw:.3f}" '
              f'stroke-dasharray="{3 / scale:.3f},{2 / scale:.3f}">'
              + "".join(f'<path d="{d}"/>' for d in proj["hidden"]) + "</g>")
        g += (f'<g stroke="#111" stroke-width="{sw:.3f}">'
              + "".join(f'<path d="{d}"/>' for d in proj["visible"]) + "</g>")
        g += "</g>"
        self.items.append(g)
        return left, top, left + dw, top + dh

    # Below this many pixels a figure no longer fits between its own
    # arrowheads; the arrows go outside and the figure beside them, the way
    # a draughtsman dimensions a sheet edge.
    NARROW = 46

    def dim_h(self, left, right, y, value):
        """A horizontal dimension under a view: extension ticks, a line with
        arrowheads, the figure above it. `value` in mm, two decimals."""
        self.line(left, y - 12, left, y + 4, w=0.8)
        self.line(right, y - 12, right, y + 4, w=0.8)
        label = f"{value:.2f}"
        if right - left >= self.NARROW:
            self.line(left, y, right, y, w=0.9)
            self.arrow(left, y, -1, 0)
            self.arrow(right, y, 1, 0)
            self.text((left + right) / 2, y - 5, label, size=13,
                      anchor="middle")
        else:
            self.line(left - 16, y, right + 16, y, w=0.9)
            self.arrow(left, y, 1, 0)
            self.arrow(right, y, -1, 0)
            self.text(right + 22, y + 4, label, size=13, anchor="start")

    def dim_v(self, top, bottom, x, value):
        """A vertical dimension beside a view, the figure reading upward."""
        self.line(x - 12, top, x + 4, top, w=0.8)
        self.line(x - 12, bottom, x + 4, bottom, w=0.8)
        label = f"{value:.2f}"
        if bottom - top >= self.NARROW:
            self.line(x, top, x, bottom, w=0.9)
            self.arrow(x, top, 0, -1)
            self.arrow(x, bottom, 0, 1)
            self.text(x + 14, (top + bottom) / 2 + 4, label, size=13,
                      anchor="middle", rotate=-90)
        else:
            self.line(x, top - 16, x, bottom + 16, w=0.9)
            self.arrow(x, top, 0, 1)
            self.arrow(x, bottom, 0, -1)
            self.text(x, bottom + 32, label, size=13, anchor="middle")

    def svg(self, height) -> str:
        return (f'<svg xmlns="http://www.w3.org/2000/svg" width="{self.W}" '
                f'height="{height:.0f}" viewBox="0 0 {self.W} {height:.0f}" '
                f'font-family="{self.FONT}"><rect width="100%" height="100%" '
                'fill="#fff"/>' + "".join(self.items) + "</svg>")


def _draw_part(sh: _Sheet, index: int, part: dict, shape, mode: str) -> None:
    """One part's band: caption and isometric on the left, front over top in
    the middle, side over the hole table on the right — the layout of the
    hand-made sheet, and of the sheet the image model used to draw."""
    M, W = sh.M, sh.W
    y0 = sh.y
    ext = _extents(shape)
    col1, col2, col3 = M, 400, 830
    vw, vh = 350, 230                 # each view's box
    # One scale for all three views of a part, so the front's width and the
    # top's width are the same line — and never larger than the box allows.
    pad = 1.0
    scale = min(vw / (ext["x"] + pad), vw / (ext["y"] + pad),
                vh / (ext["y"] + pad), vh / (ext["z"] + pad),
                6.0)                  # never blow a tiny part up past 6 px/mm
    scale = max(scale, 0.05)

    # ── caption + isometric ───────────────────────────────────────────
    sh.text(col1, y0 + 22, f"{index}) {part['name'].upper()} — 1 NOS",
            size=19, weight="bold")
    stock = (f"t≈{part['thickness_mm']:.2f} mm SHEET" if mode == "metal"
             else f"wall≈{part['thickness_mm']:.2f} mm")
    sh.text(col1 + 10, y0 + 44, stock, size=13, fill="#333")
    iso = _project(shape, (1, -1.2, 1))
    if iso:
        xmin, xmax, ymin, ymax = iso["bb"]
        s_iso = min(300 / max(xmax - xmin, 1e-6), 240 / max(ymax - ymin, 1e-6),
                    scale * 0.85)
        sh.view(iso, col1, y0 + 60, s_iso, 320, 250)

    # ── front over top ────────────────────────────────────────────────
    y_front = y0 + 30
    y_top = y_front + vh + 70
    y_side = y_front
    boxes = {"FRONT VIEW": (col2, y_front), "TOP VIEW": (col2, y_top),
             "SIDE VIEW": (col3, y_side)}
    bottom_used = y_top + vh + 40
    for title, n, xdir, (ax_w, ax_h) in _VIEWS:
        bx, by = boxes[title]
        sh.text(bx + vw / 2, by - 8, title, size=13, anchor="middle",
                weight="bold", fill="#222")
        proj = _project(shape, n, xdir)
        if not proj:
            sh.text(bx + vw / 2, by + vh / 2, "(view could not be drawn)",
                    size=12, anchor="middle", fill="#8a9098")
            continue
        left, top, right, bottom = sh.view(proj, bx, by, scale, vw, vh)
        sh.dim_h(left, right, bottom + 24, ext[ax_w])
        sh.dim_v(top, bottom, right + 26, ext[ax_h])

    # ── hole table under the side view ────────────────────────────────
    ty = y_side + vh + 70
    tx = col3 + 40
    rows = part["holes"]
    sh.text(tx, ty, f"HOLES ({part['name'].upper()})", size=13, weight="bold")
    if rows:
        for k, h in enumerate(rows):
            sh.text(tx, ty + 22 + k * 19, f"Ø{h['dia_mm']:g} x {h['count']}",
                    size=13)
        table_h = 30 + len(rows) * 19
    else:
        sh.text(tx, ty + 22, "no holes", size=13, fill="#555")
        table_h = 50
    sh.rect(tx - 12, ty - 18, 220, table_h, stroke="#777")
    # weight line, under the caption side
    sh.text(col1 + 10, y0 + 330, f"volume {part['volume_cm3']:.2f} cm3 · "
            f"{_weights(part['volume_cm3'], mode)}", size=11.5, fill="#444")

    band_bottom = max(bottom_used, ty + table_h + 10, y0 + 345)
    sh.line(M, band_bottom, W - M, band_bottom, w=1.2, color="#222")
    sh.y = band_bottom + 22


def _title_block(sh: _Sheet, report: dict) -> None:
    M, W = sh.M, sh.W
    y = sh.y
    sh.text(M, y + 16, "NOTES:", size=12.5, weight="bold")
    notes = ["All dimensions are in millimetres (mm), to two decimals.",
             "Sizes are of the FORMED part as modelled; a bent sheet's flat "
             "pattern is not shown.",
             "Thickness is an estimate (2 x volume / surface area).",
             "Hole counts are distinct hole positions; a slot reads as two."]
    for k, n in enumerate(notes):
        sh.text(M + 14, y + 36 + k * 18, f"{k + 1}. {n}", size=12, fill="#222")
    y += 36 + len(notes) * 18 + 16
    rows = [("JOB NAME:", report["file"], "DRAWN BY:", "Prism"),
            ("MATERIAL:", "CRC SHEET" if report["mode"] == "metal"
             else "MOULDED PLASTIC", "DATE:",
             _dt.date.today().strftime("%d-%m-%Y")),
            ("SCALE:", "NTS", "REMARKS:", "Measured offline by Prism — the "
                                         "STEP file never left this machine")]
    rh = 34
    sh.rect(M, y, W - 2 * M, rh * len(rows), sw=1.4)
    mid = (W + M) / 2 - 60
    for k, (a, b, c, d) in enumerate(rows):
        yy = y + k * rh
        if k:
            sh.line(M, yy, W - M, yy, w=0.8)
        sh.line(mid, yy, mid, yy + rh, w=0.8)
        sh.text(M + 12, yy + 22, a, size=11.5, fill="#444")
        sh.text(M + 140, yy + 22, b, size=15)
        sh.text(mid + 12, yy + 22, c, size=11.5, fill="#444")
        sh.text(mid + 120, yy + 22, d, size=14)
    sh.y = y + rh * len(rows) + sh.M


def sheet_svg(report: dict) -> str:
    """The whole dimensioned sheet as one SVG string — every part's three
    views with its overall sizes on dimension lines, hole table, notes and
    title block. Pure geometry and the measured figures; no AI."""
    sh = _Sheet()
    o = report["overall_mm"]
    sh.text(sh.M, sh.y + 8, f"{report['file']} — measured drawing sheet",
            size=20, weight="bold")
    sh.text(sh.M, sh.y + 30,
            f"{report['mode']} moulding · assembly overall {o[0]:.2f} x "
            f"{o[1]:.2f} x {o[2]:.2f} mm · {len(report['parts'])} part(s) · "
            f"measured offline by Prism, "
            f"{_dt.date.today().strftime('%d-%m-%Y')}",
            size=13, fill="#555")
    sh.y += 56
    sh.line(sh.M, sh.y, sh.W - sh.M, sh.y, w=1.2, color="#222")
    sh.y += 22
    shapes = dict((name, s) for name, s in report.get("_shapes") or [])
    for i, part in enumerate(report["parts"], 1):
        shape = shapes.get(part["name"])
        if shape is None:
            sh.text(sh.M, sh.y + 22, f"{i}) {part['name'].upper()} — "
                    f"{_caption(part, report['mode'])}", size=16, weight="bold")
            sh.text(sh.M + 10, sh.y + 44, "(no geometry to draw)", size=12,
                    fill="#8a9098")
            sh.y += 70
            continue
        _draw_part(sh, i, part, shape, report["mode"])
    _title_block(sh, report)
    return sh.svg(sh.y)


def render_sheet(report: dict, out_dir: str) -> dict:
    """'<model> - drawing sheet.html' / '.svg' (+ '.png' when a browser
    engine is present): the dimensioned sheet, drawn here from the
    geometry. The per-part isometric views also land beside it as
    '<model> - view <part>.svg', for anyone who wants one part on its own."""
    os.makedirs(out_dir, exist_ok=True)
    stem = _stem(report)
    out = names(stem)
    shapes = dict((name, s) for name, s in report.get("_shapes") or [])
    for i, part in enumerate(report["parts"], 1):
        shape = shapes.get(part["name"])
        if shape is not None:
            _svg_of(shape, os.path.join(out_dir, view_name(stem, part["name"], i)))

    svg = sheet_svg(report)
    svg_path = os.path.join(out_dir, out["svg"])
    with open(svg_path, "w", encoding="utf-8") as f:
        f.write(svg)
    # The captions the older sheet carried, kept as plain text under the
    # drawing so the page still READS without the picture (and so a search
    # for "101.00 x 93.00 x 71.00" finds the sheet).
    captions = "".join(
        f"<li><b>{i}) {p['name']}</b> — {_caption(p, report['mode'])}</li>"
        for i, p in enumerate(report["parts"], 1))
    html = (
        "<!doctype html><meta charset='utf-8'>"
        f"<title>{report['file']}</title>"
        "<style>body{margin:0;background:#fff;font:13px/1.4 -apple-system,"
        "'Segoe UI',sans-serif;color:#16181a}svg{display:block}"
        ".text{max-width:1240px;padding:18px 44px 30px}"
        ".text li{margin:2px 0}.note{color:#8a6d1f;font-size:12.5px}</style>"
        + svg
        + f"<div class='text'><ul>{captions}</ul>"
        + "".join(f"<p class='note'>! {w}</p>" for w in report["warnings"])
        + "</div>")
    html_path = os.path.join(out_dir, out["html"])
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    png_path = ""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": _Sheet.W, "height": 900})
            page.goto("file://" + _quote(html_path), wait_until="load")
            png_path = os.path.join(out_dir, out["png"])
            page.screenshot(path=png_path, full_page=True)
            browser.close()
    except Exception:                               # noqa: BLE001
        png_path = ""   # the HTML and SVG stand on their own
    return {"html": html_path, "svg": svg_path, "png": png_path}


def auto_brief(report: dict) -> str:
    """The prompt /step-auto hands the image agent.

    The security model lives IN the prompt: only these measured figures and
    Prism's own plain render travel to the tool — the customer's STEP file
    stays on this machine, and the brief says so out loud, so the agent
    neither asks for the model nor invents a number to fill a gap.
    """
    o = report["overall_mm"]
    lines = [
        "Draw ONE professional sheet-metal fabrication drawing sheet as a "
        "single portrait image, in the plain style of a manufacturer's "
        "hand-made dimension sheet: black line-work on white, one labelled "
        "view per part, dimension lines with arrowheads for length, width "
        "and height, hole callouts written as diameter x count, and a small "
        "title block at the bottom.",
        "",
        "Every figure below was MEASURED OFFLINE by Prism from the "
        "customer's CAD model. The model itself is confidential and is not "
        "shared — the attached image is Prism's own plain render of the "
        "parts, for view reference only.",
        "",
        f"Job: {report['file']} · {report['mode']} moulding · "
        f"assembly overall {o[0]:.2f} x {o[1]:.2f} x {o[2]:.2f} mm · "
        f"{len(report['parts'])} part(s)",
        "",
    ]
    for i, part in enumerate(report["parts"], 1):
        L, W, H = part["size_mm"]
        lines.append(f"{i}) {part['name']} — {L:.2f} x {W:.2f} x {H:.2f} mm "
                     f"· sheet t≈{part['thickness_mm']:.2f} mm · 1 nos")
        if part["holes"]:
            lines.append("   holes: " + " · ".join(
                f"Ø{h['dia_mm']:g} x {h['count']}" for h in part["holes"]))
    lines += [
        "",
        "Title block: job name, material "
        + ("CRC sheet" if report["mode"] == "metal" else "moulded plastic")
        + ", scale NTS, today's date, and the line "
          "'Measured offline by Prism'.",
        "Rules: use EXACTLY the numbers above — do not round, convert or "
        "invent any dimension, and do not add parts or holes that are not "
        "listed. Label every part with its name.",
        "Never label any view 'flat pattern' or 'developed' — no unfolded "
        "flat has been computed, and a fabricator would cut from it. Views "
        "are of the FORMED part only. Do not state any tolerance, grade or "
        "finish that is not written above; general notes may only say the "
        "dimensions are in millimetres and hole positions are indicative.",
    ]
    return "\n".join(lines)


# ── /step-ask: a question → Groq's advice → an agent's plan → applied ───────
# The customer's model still never leaves this machine. Groq gets measured
# numbers and the question; the browser agent gets those plus Groq's advice;
# and the geometry edits themselves happen HERE, in cadquery, on a copy.

PLAN_OPS = ("enlarge_hole", "scale")


def _part_lines(report: dict) -> str:
    lines = []
    for i, part in enumerate(report["parts"], 1):
        L, W, H = part["size_mm"]
        lines.append(f"{i}) {part['name']} — {L:.2f} x {W:.2f} x {H:.2f} mm "
                     f"· wall/sheet t≈{part['thickness_mm']:.2f} mm · "
                     f"volume {part['volume_cm3']:.2f} cm3")
        if part["holes"]:
            lines.append("   holes: " + " · ".join(
                f"Ø{h['dia_mm']:g} x {h['count']}" for h in part["holes"]))
    return "\n".join(lines)


def ask_prompt(report: dict, question: str) -> str:
    """What Groq is asked. Numbers and the question — never the model."""
    o = report["overall_mm"]
    return (
        f"You are advising a {report['mode']} moulding shop on a customer's "
        "part. The CAD model is confidential and cannot be shown to you — "
        "everything known about it was measured offline and is below.\n\n"
        f"Job: {report['file']} · assembly overall "
        f"{o[0]:.2f} x {o[1]:.2f} x {o[2]:.2f} mm · "
        f"{len(report['parts'])} part(s)\n"
        f"{_part_lines(report)}\n\n"
        f"The customer asks: {question}\n\n"
        "Give short, numbered, practical suggestions grounded ONLY in the "
        "figures above — do not invent features you cannot see. Where a "
        "suggestion is a hole size change or an overall scale change, state "
        "it precisely: which part, current Ø, new Ø (or scale factor), and "
        "why. Mark anything that would need the customer's designer (ribs, "
        "draft, wall changes) as their decision, not ours.")


def plan_prompt(report: dict, question: str, suggestions: str) -> str:
    """What the reviewing agent is asked: turn the advice into a strict
    machine plan of ONLY the operations Prism can execute locally."""
    return (
        "You are the reviewing engineer. Below are offline measurements of "
        "a confidential CAD model (the model itself is not shared) and a "
        "first advisor's suggestions. Decide which changes are right, then "
        "answer with ONE JSON object and nothing else.\n\n"
        f"MEASURED ({report['mode']} moulding, {report['file']}):\n"
        f"{_part_lines(report)}\n\n"
        f"THE CUSTOMER ASKED: {question}\n\n"
        f"FIRST ADVISOR SAID:\n{suggestions}\n\n"
        "Prism can execute exactly two operations on the model, locally:\n"
        '  {"op": "enlarge_hole", "part": "<part name or all>", '
        '"dia_mm": <current>, "new_dia_mm": <bigger>, "why": "..."}\n'
        '  {"op": "scale", "part": "<part name or all>", '
        '"factor": <0.2..5>, "why": "..."}\n\n'
        "Answer format:\n"
        '{"changes": [ ...only the two ops above, only if truly right... ],\n'
        ' "advice":  [ "every other worthwhile suggestion, as a sentence" ]}\n\n'
        "Rules: use only part names and hole diameters that appear in the "
        "measurements. A hole can only be enlarged, never shrunk. When no "
        "executable change is justified, return an empty changes list — an "
        "honest empty list beats an invented edit.")


def _valid_change(ch) -> dict | None:
    if not isinstance(ch, dict):
        return None
    part = str(ch.get("part") or "all").strip() or "all"
    why = str(ch.get("why") or "")[:240]
    try:
        if ch.get("op") == "enlarge_hole":
            dia, new = float(ch["dia_mm"]), float(ch["new_dia_mm"])
            if not 0 < dia < new:
                return None
            return {"op": "enlarge_hole", "part": part, "dia_mm": round(dia, 2),
                    "new_dia_mm": round(new, 2), "why": why}
        if ch.get("op") == "scale":
            factor = float(ch["factor"])
            if not 0.2 <= factor <= 5 or factor == 1:
                return None
            return {"op": "scale", "part": part,
                    "factor": round(factor, 4), "why": why}
    except (KeyError, TypeError, ValueError):
        return None
    return None


def parse_plan(texts: list[str]) -> tuple[dict | None, str]:
    """Newest capture that parses — the tab also holds the prompt Prism
    typed, which carries the example schema inside it."""
    import json
    for t in reversed([t for t in texts if t and t.strip()]):
        s, e = t.find("{"), t.rfind("}") + 1
        if s == -1 or e <= s:
            continue
        try:
            raw = json.loads(t[s:e])
        except ValueError:
            continue
        if not isinstance(raw, dict) or not (
                "changes" in raw or "advice" in raw):
            continue
        changes = [c for c in map(_valid_change, raw.get("changes") or [])
                   if c]
        advice = [str(a).strip() for a in (raw.get("advice") or [])
                  if str(a).strip()][:12]
        return {"changes": changes, "advice": advice}, ""
    return None, "The agent returned no JSON plan Prism could read."


def _enlarged(shape, dia: float, new_dia: float):
    """Cut a bigger cylinder along every existing axis of the Ø`dia` holes.
    Enlarge only — shrinking would mean adding material, which a boolean
    cut cannot do and _valid_change refuses upstream."""
    bb = shape.BoundingBox()
    span = (bb.xlen ** 2 + bb.ylen ** 2 + bb.zlen ** 2) ** 0.5 or 1.0
    seen, cutters = set(), []
    for f in shape.Faces():
        if f.geomType() != "CYLINDER":
            continue
        cyl = BRepAdaptor_Surface(TopoDS.Face_s(f.wrapped)).Cylinder()
        if abs(2 * cyl.Radius() - dia) > 0.05:
            continue
        ax = cyl.Axis()
        loc, d = ax.Location(), ax.Direction()
        key = (round(loc.X(), 1), round(loc.Y(), 1), round(loc.Z(), 1))
        if key in seen:
            continue
        seen.add(key)
        start = _cq.Vector(loc.X(), loc.Y(), loc.Z()) - \
            _cq.Vector(d.X(), d.Y(), d.Z()) * span
        cutters.append(_cq.Solid.makeCylinder(
            new_dia / 2, 2 * span, pnt=start,
            dir=_cq.Vector(d.X(), d.Y(), d.Z())))
    for c in cutters:
        shape = shape.cut(c)
    return shape, len(seen)


def _scaled(shape, factor: float):
    """gp_Trsf, not transformGeometry: the general transform rewrites every
    cylinder as a b-spline, and a hole that is no longer a CYLINDER face
    vanishes from _holes_of — the re-measure would deny holes that exist."""
    from OCP.BRepBuilderAPI import BRepBuilderAPI_Transform
    from OCP.gp import gp_Pnt, gp_Trsf
    t = gp_Trsf()
    t.SetScale(gp_Pnt(0, 0, 0), factor)
    return _cq.Shape.cast(
        BRepBuilderAPI_Transform(shape.wrapped, t, True).Shape())


def apply_plan(path: str, plan: dict, out_path: str) -> dict:
    """Execute the validated plan on a COPY; the original file is never
    written. Returns {"out": path, "log": [human lines]}."""
    parts = load_parts(path)
    by_name = dict(parts)
    log = []
    for ch in plan.get("changes") or []:
        names = ([n for n, _s in parts] if ch["part"] == "all"
                 else [n for n, _s in parts if n == ch["part"]])
        if not names:
            log.append(f"! no part called '{ch['part']}' — skipped")
            continue
        for name in names:
            if ch["op"] == "enlarge_hole":
                shape, n = _enlarged(by_name[name],
                                     ch["dia_mm"], ch["new_dia_mm"])
                if n:
                    by_name[name] = shape
                    log.append(f"Ø{ch['dia_mm']:g} → Ø{ch['new_dia_mm']:g} "
                               f"on {name}: {n} hole(s) enlarged")
                elif ch["part"] != "all":
                    log.append(f"! no Ø{ch['dia_mm']:g} hole on {name} "
                               "— skipped")
            elif ch["op"] == "scale":
                by_name[name] = _scaled(by_name[name], ch["factor"])
                log.append(f"{name} scaled x {ch['factor']:g}")
    if not any(not line.startswith("!") for line in log):
        raise StepError("Nothing in the plan could be applied — "
                        "the model was not written.")
    asm = _cq.Assembly()
    for name, _s in parts:
        asm.add(by_name[name], name=name)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    asm.export(out_path)
    return {"out": out_path, "log": log}


def predicted_parts(report: dict, plan: dict) -> list[dict]:
    """What the parts SHOULD measure after the plan — arithmetic on the
    measured figures, for the review page. The real answer still comes from
    re-measuring the built file; this is the promise the user approves."""
    import copy
    parts = copy.deepcopy(report["parts"])
    for ch in plan.get("changes") or []:
        for p in parts:
            if ch["part"] not in ("all", p["name"]):
                continue
            if ch["op"] == "enlarge_hole":
                for h in p["holes"]:
                    if abs(h["dia_mm"] - ch["dia_mm"]) <= 0.05:
                        h["dia_mm"] = ch["new_dia_mm"]
                merged: dict[float, int] = {}
                for h in p["holes"]:
                    merged[h["dia_mm"]] = merged.get(h["dia_mm"], 0) + h["count"]
                p["holes"] = [{"dia_mm": d, "count": c}
                              for d, c in sorted(merged.items())]
            elif ch["op"] == "scale":
                f = ch["factor"]
                p["size_mm"] = tuple(round(v * f, 2) for v in p["size_mm"])
                p["thickness_mm"] = round(p["thickness_mm"] * f, 2)
                p["volume_cm3"] = round(p["volume_cm3"] * f ** 3, 2)
                p["holes"] = [{"dia_mm": round(h["dia_mm"] * f, 2),
                               "count": h["count"]} for h in p["holes"]]
    return parts


def _holes_str(part: dict) -> str:
    return " · ".join(f"Ø{h['dia_mm']:g} x {h['count']}"
                      for h in part["holes"]) or "—"


def review_html(report: dict, plan: dict, out_dir: str,
                question: str = "", after: dict | None = None) -> str:
    """The approval page: every dimension before and after, side by side,
    with the drawing image — written BEFORE anything is built, so the user
    confirms against what they can see, not against terminal text. Called
    again with `after` (the re-measured report) once modified.step exists,
    so the same page becomes the record of what was actually done."""
    import html as _html

    before = report["parts"]
    shown = ([{"name": p["name"], "size_mm": p["size_mm"],
               "thickness_mm": p["thickness_mm"],
               "volume_cm3": p["volume_cm3"], "holes": p["holes"]}
              for p in after["parts"]] if after
             else predicted_parts(report, plan))
    by_after = {p["name"]: p for p in shown}

    if after:
        banner = (f"<div class='banner'>{_html.escape(names(report)['modified'])}"
                  " is BUILT — the After column is re-measured from the new "
                  "file, not predicted.</div>")
    else:
        banner = ("<div class='banner'>Nothing is built yet. This page is "
                  "the plan — go back to the terminal and answer Y to write "
                  f"{_html.escape(names(report)['modified'])}, or N to stop "
                  "here.</div>")

    def fmt(size):
        return f"{size[0]:.2f} x {size[1]:.2f} x {size[2]:.2f}"

    change_rows = []
    for ch in plan.get("changes") or []:
        what = (f"Hole Ø{ch['dia_mm']:g} → Ø{ch['new_dia_mm']:g}"
                if ch["op"] == "enlarge_hole"
                else f"Scale x {ch['factor']:g}")
        change_rows.append(
            f"<tr><td>{_html.escape(ch['part'])}</td>"
            f"<td>{_html.escape(what)}</td>"
            f"<td>{_html.escape(ch.get('why') or '')}</td></tr>")

    part_rows = []
    for p in before:
        q = by_after.get(p["name"], p)
        changed_size = p["size_mm"] != q["size_mm"]
        changed_holes = _holes_str(p) != _holes_str(q)
        mark = " class='changed'"
        part_rows.append(
            "<tr>"
            f"<td>{_html.escape(p['name'])}</td>"
            f"<td>{fmt(p['size_mm'])}</td>"
            f"<td{mark if changed_size else ''}>{fmt(q['size_mm'])}</td>"
            f"<td>{_holes_str(p)}</td>"
            f"<td{mark if changed_holes else ''}>{_holes_str(q)}</td>"
            f"<td>{p['thickness_mm']:.2f} → {q['thickness_mm']:.2f}</td>"
            "</tr>")

    advice = "".join(f"<li>{_html.escape(a)}</li>"
                     for a in plan.get("advice") or [])
    out = names(report)
    imgs = []
    if os.path.exists(os.path.join(out_dir, out["png"])):
        imgs.append(("The part as received", out["png"]))
    if after and os.path.exists(os.path.join(out_dir, out["png_after"])):
        imgs.append(("The modified model — drawn from the BUILT file",
                     out["png_after"]))
    img = "".join(f"<h2>{t}</h2><img src='{_quote(f)}'>" for t, f in imgs)

    page = (
        "<!doctype html><meta charset='utf-8'>"
        f"<title>{_html.escape(report['file'])} — change review</title>"
        "<style>body{font-family:-apple-system,Segoe UI,sans-serif;margin:32px "
        "auto;max-width:1000px;color:#1c2733;background:#fafbfc;padding:0 16px}"
        "h1{font-size:22px}h2{font-size:15px;margin-top:28px}"
        "table{border-collapse:collapse;width:100%;font-size:13.5px}"
        "th,td{border:1px solid #d7dade;padding:7px 10px;text-align:left}"
        "th{background:#eef1f4}td.changed{background:#e7f6ec;font-weight:600}"
        ".banner{background:#fff4d6;border:1px solid #e3c96e;padding:12px "
        "16px;border-radius:8px;margin:14px 0;font-weight:600}"
        ".banner.built{background:#e7f6ec;border-color:#7cc796}"
        ".q{color:#41586e}img{max-width:100%;border:1px solid #d7dade;"
        "border-radius:8px}.note{color:#8a6d1f;font-size:12.5px}</style>"
        f"<h1>{_html.escape(report['file'])} — change review</h1>"
        f"<p class='q'>Asked: {_html.escape(question)}</p>"
        f"{banner}"
        "<h2>Changes Prism will make"
        + (" (made)" if after else "") + "</h2>"
        "<table><tr><th>Part</th><th>Change</th><th>Why</th></tr>"
        + "".join(change_rows) + "</table>"
        "<h2>Every dimension, before → after</h2>"
        "<table><tr><th>Part</th><th>Size before (mm)</th>"
        f"<th>Size {'after' if after else 'after (predicted)'} (mm)</th>"
        "<th>Holes before</th>"
        f"<th>Holes {'after' if after else 'after (predicted)'}</th>"
        "<th>t (mm)</th></tr>"
        + "".join(part_rows) + "</table>"
        + (f"<h2>For the designer (not applied)</h2><ul>{advice}</ul>"
           if advice else "")
        + img
        + "<p class='note'>Measured offline by Prism — the STEP file never "
          "left this machine; the edits run locally on a copy.</p>")

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, out["review"])
    with open(path, "w", encoding="utf-8") as f:
        f.write(page)
    return path
