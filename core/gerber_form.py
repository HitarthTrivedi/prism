"""Prism — fill the client's own quotation form from a measured Gerber job.

The client already has a form. Fine Circuits (Manjusar GIDC) hands their
estimator F-SAL-01 — an Excel quotation sheet with a cell for every figure:
Board X/Y, Array X/Y, Pcs/Array, Min Line, Min Space, Smallest Hole, Pitch,
SMT pad, a drill table summed by the form's own =SUM() — and what they want
from Prism is not a new report format, it is THEIR form with the measured
numbers already in it.

How filling works — labels, not coordinates. The template is scanned for
cells whose text matches a known label ("Board X", "Min Line", "Smallest
Hole", "Pcs / Array"…) and the measured value is written into the cell
immediately to the label's right — which is how these forms are invariably
laid out. Coordinates would pin Prism to one revision of one client's file;
labels survive the client inserting a row, and they make the same code fill
the NEXT client's differently-drawn form too.

Two rules keep the client's form theirs:

  · A cell that holds a formula is never overwritten. The form's own
    arithmetic (=SUM(B30:B51), =B30*$B$13) keeps working on the values
    Prism puts in — their totals stay THEIR totals.
  · A label Prism has no measurement for is left exactly as drawn. An
    empty cell on their form is a question for their estimator, not a
    place for a guess.

The drill table is the one structural fill: the row below "HOLE SIZE" gets
one row per measured tool — diameter in mm, hits per array — and any
leftover pre-printed placeholder rows are zeroed so the form's SUM counts
only real drills.

Figures are written to 2 decimal places, same as every Gerber surface.
"""
from __future__ import annotations

import datetime as _dt
import os
import re

try:
    import openpyxl
    HAVE_XLSX = True
except Exception:                                   # pragma: no cover
    HAVE_XLSX = False


class FormError(Exception):
    """A sentence for the user, not a stack trace."""


def _round2(v):
    return round(float(v), 2)


def _smt_lw(answers: dict) -> tuple:
    """(length, width) of the smallest SMT pad, longest side first."""
    text = answers.get("min_smt_pad") or ""
    m = re.match(r"([\d.]+)\s*x\s*([\d.]+)", text)
    if not m:
        return None, None
    a, b = float(m.group(1)), float(m.group(2))
    return max(a, b), min(a, b)


def _size(answers: dict, key: str, idx: int):
    pair = answers.get(key)
    return _round2(pair[idx]) if pair and pair[idx] else None


# The unit every measured LENGTH is written in. Prism measures in mm; a fab
# whose form (or customer) works in thou wants the same cells in mil or
# inch. Counts, layer numbers and text never convert — only lengths do.
UNITS = ("mm", "inch", "mil")
_PER_MM = {"mm": 1.0, "inch": 1 / 25.4, "mil": 1000 / 25.4}
_DECIMALS = {"mm": 2, "inch": 4, "mil": 2}


def _in_units(v_mm, units: str):
    """One length, mm → the chosen unit, at that unit's honest precision
    (0.25 mm is 0.0098 inch — two decimals would round a track width to a
    different track)."""
    return round(float(v_mm) * _PER_MM[units], _DECIMALS[units])


# ── the extras a job knows beyond the nine headline figures ─────────────────
# Everything here is read off the job the measurement already built — file
# roles, drill flags, and the CAM tool's own text reports. A fact the job
# does not carry stays None, and the form cell stays blank: an empty cell on
# the client's form is a question for their estimator, not a place for a
# guess.

_FINISHES = re.compile(
    r"\b(ENIG|HASL|HAL(?:\s*LF)?|OSP|IMMERSION\s+(?:GOLD|SILVER|TIN))\b",
    re.IGNORECASE)
_MATERIALS = re.compile(
    r"\b(FR-?4|CEM-?[13]|POLYIMIDE|METAL\s*CORE|ALUMINI?UM)\b", re.IGNORECASE)
_THICKNESS = re.compile(
    r"THICK\w*\W{0,15}?(\d(?:\.\d{1,2})?)\s*MM", re.IGNORECASE)
_COPPER_OZ = re.compile(r"(\d(?:\.\d)?)\s*OZ", re.IGNORECASE)
_INNER_OZ = re.compile(
    r"(?:I/?L|INNER)\W{0,20}?(\d(?:\.\d)?)\s*OZ", re.IGNORECASE)
_SCORING = re.compile(r"\bV[- ]?CUT\b|\bSCOR(?:E|ING)\b", re.IGNORECASE)
_MASK_COLOR = re.compile(
    r"(GREEN|BLUE|BLACK|WHITE|RED|YELLOW|PURPLE|MATT[E]?\s+GREEN)"
    r"\W{0,20}(?:SOLDER\s*)?MASK"
    r"|(?:SOLDER\s*)?MASK\W{0,20}"
    r"(GREEN|BLUE|BLACK|WHITE|RED|YELLOW|PURPLE|MATT[E]?\s+GREEN)",
    re.IGNORECASE)
_UL_CODE = re.compile(r"\bUL\W{0,15}(E\s?\d{5,6})\b", re.IGNORECASE)


def _report_text(job: dict, cap: int = 200_000) -> str:
    """Every CAM report / stackup note in the job, as one searchable text."""
    chunks = []
    for f in job.get("files") or []:
        name = (f.get("name") or "").lower()
        if f.get("role") in ("report", "rules") or "stackup" in name:
            try:
                with open(f["path"], "r", errors="replace") as fh:
                    chunks.append(fh.read(cap))
            except Exception:                           # noqa: BLE001
                pass
    return "\n".join(chunks)


def _sides_word(roles: set, top: str, bottom: str):
    has_top, has_bottom = top in roles, bottom in roles
    if has_top and has_bottom:
        return "BOTH"
    if has_top:
        return "TOP"
    if has_bottom:
        return "BOTTOM"
    return None


def _job_extras(job: dict) -> dict:
    files = job.get("files") or []
    roles = {f.get("role") for f in files}
    out = {
        "sm_sides": _sides_word(roles, "mask_top", "mask_bottom"),
        "legend_sides": _sides_word(roles, "silk_top", "silk_bottom"),
    }

    # Which side carries SMT pads — from the measured pad rows.
    by_name = {f.get("name"): f.get("role") for f in files}
    smt_roles = {by_name.get(r.get("name"))
                 for r in job.get("smt") or [] if r.get("count")}
    out["smt_side"] = _sides_word(smt_roles, "copper_top", "copper_bottom")

    # Slots: the drill file says so itself — G85 slot commands, or a rout
    # pass. No drill file at all means "don't know", not "no".
    drills = job.get("drills") or {}
    drill_files = [f for f in files if f.get("role") == "drill"]
    if drill_files:
        slot = bool(drills.get("is_rout"))
        if not slot:
            for f in drill_files:
                try:
                    with open(f["path"], "r", errors="replace") as fh:
                        slot = "G85" in fh.read(2_000_000)
                except Exception:                       # noqa: BLE001
                    pass
                if slot:
                    break
        out["slots"] = "YES" if slot else "NO"

    # Stackup facts mined from the CAM tool's own text reports — taken only
    # when the report states them, in the report's own words.
    text = _report_text(job)
    if text:
        m = _MATERIALS.search(text)
        if m:
            out["material_type"] = m.group(1).upper().replace("FR4", "FR-4")
        m = _THICKNESS.search(text)
        if m:
            out["material_thickness"] = f"{m.group(1)} mm"
        m = _COPPER_OZ.search(text)
        if m:
            out["copper_weight"] = f"{m.group(1)} oz"
        m = _INNER_OZ.search(text)
        if m:
            out["inner_copper_weight"] = f"{m.group(1)} oz"
        m = _FINISHES.search(text)
        if m:
            out["final_finish"] = m.group(1).upper()
        if _SCORING.search(text):
            out["scoring"] = "YES"
        m = _MASK_COLOR.search(text)
        if m:
            out["sm_color"] = (m.group(1) or m.group(2)).upper()
        m = _UL_CODE.search(text)
        if m:
            out["ul_code"] = m.group(1).replace(" ", "")
    return out


# (label regex, getter(answers, meta) -> value or None, is_length). Matched
# against the cell text lowercased with trailing ':'/'-' stripped, so
# "CUSTOMER:-" and "Board X" both read naturally. First match wins; order
# the specific before the generic. `is_length` marks the values that follow
# the chosen unit — everything else is a count or text and never converts.
_LABELS = [
    (r"^no\.?\s*layers?$|^layers$|^no\s+layer$",
     lambda a, m: a.get("layers") or None, False),
    (r"^board\s*x$", lambda a, m: _size(a, "pcb_size_mm", 0), True),
    (r"^board\s*y$", lambda a, m: _size(a, "pcb_size_mm", 1), True),
    (r"^array\s*x$", lambda a, m: _size(a, "array_size_mm", 0), True),
    (r"^array\s*y$", lambda a, m: _size(a, "array_size_mm", 1), True),
    (r"^pcs\s*/\s*array$|^pcbs?\s*(in|per)\s*(the\s*)?array$",
     lambda a, m: a.get("pcbs_per_array") or None, False),
    (r"^min\.?\s*line$|^min\.?\s*track(\s*width)?$",
     lambda a, m: _round2(a["min_track_width_mm"])
     if a.get("min_track_width_mm") else None, True),
    (r"^min\.?\s*space$|^min\.?\s*(track\s*)?spacing$",
     lambda a, m: _round2(a["min_track_spacing_mm"])
     if a.get("min_track_spacing_mm") else None, True),
    (r"^smallest\s*hole$|^min\.?\s*drill(\s*size)?$",
     lambda a, m: _round2(a["min_drill_mm"])
     if a.get("min_drill_mm") else None, True),
    (r"^(min\.?\s*)?pitch$",
     lambda a, m: _round2(a["min_pitch_mm"])
     if a.get("min_pitch_mm") else None, True),
    (r"^min\.?\s*smt\s*length$", lambda a, m: _smt_lw(a)[0], True),
    (r"^min\.?\s*smt\s*width$", lambda a, m: _smt_lw(a)[1], True),
    (r"^min\.?\s*annular\s*ring$",
     lambda a, m: _round2(a["min_annular_ring_mm"])
     if a.get("min_annular_ring_mm") else None, True),
    (r"^no\.?\s*of\s*tools$", lambda a, m: a.get("drill_tools"), False),
    (r"^smt\s*sides?$", lambda a, m: a.get("smt_side"), False),
    (r"^sm\s*sides?$", lambda a, m: a.get("sm_sides"), False),
    (r"^legend\s*sides?$", lambda a, m: a.get("legend_sides"), False),
    (r"^slots?$", lambda a, m: a.get("slots"), False),
    # S&R is Step & Repeat — the measured array arrangement (e.g. "2 x 3").
    (r"^s\s*&\s*r$", lambda a, m: a.get("array_grid"), False),
    (r"^i/?l\s*wt\.?(\s*finish)?$",
     lambda a, m: a.get("inner_copper_weight"), False),
    (r"^sm\s*colou?r$", lambda a, m: a.get("sm_color"), False),
    (r"^ul\s*code$", lambda a, m: a.get("ul_code"), False),
    (r"^material\s*type$", lambda a, m: a.get("material_type"), False),
    (r"^mat\.?\s*thickness$",
     lambda a, m: a.get("material_thickness"), False),
    (r"^cu\.?\s*wt\.?(\s*finish)?$",
     lambda a, m: a.get("copper_weight"), False),
    (r"^final\s*finish$", lambda a, m: a.get("final_finish"), False),
    (r"^scoring$", lambda a, m: a.get("scoring"), False),
    (r"^customer$", lambda a, m: m.get("customer") or None, False),
    (r"^part\s*no\.?$", lambda a, m: m.get("part") or None, False),
    (r"^date$", lambda a, m: m.get("date")
     or _dt.date.today().strftime("%d-%m-%Y"), False),
]
_COMPILED = [(re.compile(rx, re.IGNORECASE), get, linear)
             for rx, get, linear in _LABELS]

_DRILL_HEADER = re.compile(r"hole\s*size", re.IGNORECASE)


def _label_text(value) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip().rstrip(":-").strip().lower()


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _fill_drill_table(ws, job: dict, filled: list, writes: dict,
                      units: str = "mm") -> int:
    """One row per measured tool under 'HOLE SIZE'; leftover pre-printed
    placeholder rows zeroed so the form's own SUM counts only real drills.
    Stops at the first row that mentions TOTAL — that row is the form's."""
    header = next((c for row in ws.iter_rows() for c in row
                   if isinstance(c.value, str)
                   and _DRILL_HEADER.search(c.value)), None)
    if header is None:
        return 0
    tools = sorted((t for t in (job.get("drills") or {}).get("tools", [])
                    if t.get("hits")), key=lambda t: t["dia_mm"])
    col, count_col = header.column, header.column + 1
    row = header.row + 1
    written = 0
    while row <= ws.max_row:
        size_cell = ws.cell(row=row, column=col)
        count_cell = ws.cell(row=row, column=count_col)
        if isinstance(size_cell.value, str) and "total" in size_cell.value.lower():
            break
        if written < len(tools):
            t = tools[written]
            if not _is_formula(size_cell.value):
                writes[size_cell.coordinate] = _in_units(t["dia_mm"], units)
            if not _is_formula(count_cell.value):
                writes[count_cell.coordinate] = t["hits"]
            filled.append((size_cell.coordinate, "drill",
                           f"Ø{_round2(t['dia_mm']):g} x {t['hits']}"))
            written += 1
        elif size_cell.value is not None or count_cell.value:
            # A pre-printed placeholder past the measured list — zero it so
            # the form's SUM row adds up to the real drill count.
            if not _is_formula(size_cell.value):
                writes[size_cell.coordinate] = None
            if not _is_formula(count_cell.value):
                writes[count_cell.coordinate] = 0
        row += 1
    return written


# ── writing: patch the copy, never rebuild it ────────────────────────────────
# openpyxl's save() REBUILDS the workbook, and everything it does not model
# is silently rebuilt wrong or dropped: the client's logo drawing came back
# mangled and their printer settings (the page setup their form prints with)
# vanished. The client's form is their document — photos, layout, print
# margins and all — so the filled copy starts as a byte-for-byte copy of the
# template and ONLY the sheet XML carrying the changed cells is rewritten.

_XLSX_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = ("http://schemas.openxmlformats.org/officeDocument/2006/"
           "relationships")


def _sheet_parts(zf) -> dict:
    """{sheet title: 'xl/worksheets/sheetN.xml'} via workbook.xml + rels."""
    import xml.etree.ElementTree as ET
    rels = {}
    root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    for rel in root:
        target = rel.get("Target", "").lstrip("/")
        rels[rel.get("Id")] = (target if target.startswith("xl/")
                               else "xl/" + target)
    out = {}
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    for sheet in wb.iter(f"{{{_XLSX_NS}}}sheet"):
        rid = sheet.get(f"{{{_REL_NS}}}id")
        if rid in rels:
            out[sheet.get("name")] = rels[rid]
    return out


def _cell_ref(coord: str) -> tuple[int, int]:
    m = re.match(r"([A-Z]+)(\d+)", coord)
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)), col


def _cell_body(value) -> tuple[str, str]:
    """(extra attribute, inner XML) for a cell holding `value`."""
    if value is None:
        return "", ""
    if isinstance(value, str):
        esc = (value.replace("&", "&amp;").replace("<", "&lt;")
               .replace(">", "&gt;"))
        return ' t="inlineStr"', f"<is><t>{esc}</t></is>"
    num = str(value) if isinstance(value, int) else f"{value:g}"
    return "", f"<v>{num}</v>"


def _patch_sheet_xml(xml: bytes, writes: dict) -> bytes:
    """Set (or clear) cell values by STRING surgery, never by re-serialising
    the document. An XML library round-trip renames the sheet's namespace
    prefixes (mc: became ns1:) and Excel then reports the client's file as
    corrupt — so every byte outside the touched <c> elements stays exactly
    as the client's own Excel wrote it. A cell holding a formula is left
    alone; a missing cell is inserted in column order."""
    text = xml.decode("utf-8")

    for coord, value in writes.items():
        row_n, col_n = _cell_ref(coord)
        extra, inner = _cell_body(value)
        m = re.search(rf'<c\b([^>]*?\br="{coord}"[^>]*?)(/>|>)', text)
        if m:
            attrs = re.sub(r'\s+t="[^"]*"', "", m.group(1))
            if m.group(2) == "/>":
                old_end = m.end()
                body = ""
            else:
                close = text.index("</c>", m.end()) + len("</c>")
                body = text[m.end():close - len("</c>")]
                old_end = close
            if "<f" in body:
                continue                              # their arithmetic
            new = (f"<c{attrs}{extra}/>" if not inner
                   else f"<c{attrs}{extra}>{inner}</c>")
            text = text[:m.start()] + new + text[old_end:]
            continue
        if value is None:
            continue                                  # nothing to clear
        cell = (f'<c r="{coord}"{extra}/>' if not inner
                else f'<c r="{coord}"{extra}>{inner}</c>')
        rm = re.search(rf'<row\b[^>]*?\br="{row_n}"[^>]*?(/>|>)', text)
        if rm:
            if rm.group(1) == "/>":
                text = (text[:rm.start(1)] + ">" + cell + "</row>"
                        + text[rm.end(1):])
                continue
            close = text.index("</row>", rm.end())
            pos = close
            for cm in re.finditer(r'<c\b[^>]*?\br="([A-Z]+\d+)"',
                                  text[rm.end():close]):
                if _cell_ref(cm.group(1))[1] > col_n:
                    pos = rm.end() + cm.start()
                    break
            text = text[:pos] + cell + text[pos:]
            continue
        row_xml = f'<row r="{row_n}">{cell}</row>'
        pos = None
        for rm2 in re.finditer(r'<row\b[^>]*?\br="(\d+)"', text):
            if int(rm2.group(1)) > row_n:
                pos = rm2.start()
                break
        if pos is None:
            end = text.find("</sheetData>")
            if end == -1:
                sd = re.search(r"<sheetData\s*/>", text)
                if sd is None:
                    continue
                text = (text[:sd.start()] + "<sheetData>" + row_xml
                        + "</sheetData>" + text[sd.end():])
                continue
            pos = end
        text = text[:pos] + row_xml + text[pos:]
    return text.encode("utf-8")


def _patch_xlsx(template_path: str, out_path: str,
                writes_by_sheet: dict) -> None:
    """Copy the template byte-for-byte, then rewrite only the sheet XML
    that carries changed cells — logo, drawings, print setup, styles and
    every other part of the client's file stay exactly as they made it."""
    import shutil
    import zipfile

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with zipfile.ZipFile(template_path) as zin:
        parts = _sheet_parts(zin)
        patched = {}
        for title, writes in writes_by_sheet.items():
            if writes and title in parts:
                patched[parts[title]] = _patch_sheet_xml(
                    zin.read(parts[title]), writes)
        if not patched:
            shutil.copyfile(template_path, out_path)
            return
        with zipfile.ZipFile(out_path, "w",
                             zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = patched.get(item.filename) or zin.read(item.filename)
                zout.writestr(item, data)


def fill_form(job: dict, template_path: str, out_path: str,
              meta: dict | None = None, units: str = "mm") -> dict:
    """Write a filled COPY of the client's form; the template is never
    touched, and neither is anything in the copy except the cells that
    got a measured value. Every length goes in as `units` (mm, inch or
    mil); counts and text never convert. Returns {"out", "filled":
    [(cell, label, value)…], "drill_rows": n, "units"}."""
    if not HAVE_XLSX:
        raise FormError("Filling an Excel form needs the openpyxl package.")
    if units not in UNITS:
        raise FormError(f"Units must be one of {UNITS}, not {units!r}.")
    template_path = os.path.abspath(os.path.expanduser(template_path))
    if not os.path.exists(template_path):
        raise FormError(f"No such template: {template_path}")

    # openpyxl is the READER — it finds the labels, the formulas and the
    # drill table. It never saves: writing goes through _patch_xlsx so the
    # client's photos and layout survive.
    wb = openpyxl.load_workbook(template_path)
    # The headline figures, plus everything else the job knows — file-role
    # facts and CAM-report facts land in the same lookup the labels read.
    answers = {**(job.get("answers") or {}), **_job_extras(job)}
    meta = meta or {}
    filled: list = []
    drills = 0
    writes_by_sheet: dict = {}

    for ws in wb.worksheets:
        writes = writes_by_sheet.setdefault(ws.title, {})
        for row in ws.iter_rows():
            for cell in row:
                text = _label_text(cell.value)
                if not text:
                    continue
                for rx, get, linear in _COMPILED:
                    if not rx.match(text):
                        continue
                    try:
                        value = get(answers, meta)
                    except Exception:               # noqa: BLE001
                        value = None
                    if value is None:
                        break   # measured nothing — leave their form as drawn
                    if linear and units != "mm":
                        value = _in_units(value, units)
                    target = ws.cell(row=cell.row, column=cell.column + 1)
                    if _is_formula(target.value):
                        break   # their arithmetic, never ours
                    writes[target.coordinate] = value
                    filled.append((target.coordinate, cell.value.strip(),
                                   value))
                    break
        drills += _fill_drill_table(ws, job, filled, writes, units)

    _patch_xlsx(template_path, out_path, writes_by_sheet)
    return {"out": out_path, "filled": filled, "drill_rows": drills,
            "units": units}
